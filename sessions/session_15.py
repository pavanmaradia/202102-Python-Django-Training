"""
Parse Excel file
Generators
"""

import os

import openpyxl

CURRENT_PATH = os.path.abspath(__file__)
DIR_PATH = os.path.dirname(os.path.dirname(CURRENT_PATH))
FILES_DIR_PATH = os.path.join(DIR_PATH, 'files')
XLS_PATH = os.path.join(FILES_DIR_PATH, 'test file_python.xlsx')

wb_obj = openpyxl.load_workbook(XLS_PATH)

ws = wb_obj.active


# data = []
# for i in range(1, ws.max_row+1):
#     data.append({
#         'no': ws[F'A{i}'].value,
#         'employee_id': ws[F'B{i}'].value,
#         'age': ws[F'C{i}'].value,
#         'salary': ws[F'D{i}'].value
#     })
# print(data)

# x = [1, 2, 3, 4, 5]
# y = iter(x)
# print(y.__next__())
# print(y.__next__())

# for i in x:
#     print(i)


# def my_generator(start, end, step):
#     start = start
#     while start <= end:
#         yield start
#         start += step
#
#
# for i in my_generator(1, 100000000000, 2):
#     print(i)
#

def file_parser(_file_path):
    wb_obj = openpyxl.load_workbook(XLS_PATH)
    ws = wb_obj.active
    for i in range(1, ws.max_row + 1):
        data = {
            'no': ws[F'A{i}'].value,
            'employee_id': ws[F'B{i}'].value,
            'age': ws[F'C{i}'].value,
            'salary': ws[F'D{i}'].value
        }
        yield data



age_wise = {}

"""
{
    'age_range': {
        'salary' : ['empoyee_id']
    },
    '51+': {
        124689: [100111, 100129]
    }
}
"""



for i in file_parser(XLS_PATH):
    age = i.get('age')
    salary = i.get('salary')
    employee_id = i.get('employee_id')
    if age == 'Age':
        continue
    if 21 <= age <= 30:
        _age = '21-30'
    elif 31 <= age <= 40:
        _age = '31-40'
    elif 41 <= age <= 51:
        _age = '41-50'
    else:
        _age = '51+'
    if _age not in age_wise:
        age_wise[_age] = dict()

    if salary not in age_wise[_age]:
        age_wise[_age][salary] = list()

    age_wise[_age][salary].append(employee_id)

age_group = '41-50'

highest_salary = 0
_message = ''
for age_group in ['21-30', '31-40', '41-50', '51+']:
    emp_max_salary = max(age_wise[age_group].keys())
    print(F'{age_group}: {emp_max_salary}')
    if highest_salary < emp_max_salary:
        highest_salary = emp_max_salary
        _message = F"{age_wise[age_group][emp_max_salary]} " \
                   F"gets {emp_max_salary} Salary"

print(_message)