"""
Exception Handling
    Try Except
    Try Except Else
Debugging using Pycharm debug option
"""

import os

import openpyxl


def generate_xls_path(file_name=''):
    current_path = os.path.abspath(__file__)
    dir_path = os.path.dirname(os.path.dirname(current_path))
    files_dir_path = os.path.join(dir_path, 'files')
    xls_file_path = os.path.join(files_dir_path, file_name)
    return xls_file_path
    # if os.path.exists(xls_file_path):
    #     return xls_file_path
    # else:
    #     raise Exception(FileNotFoundError)

def parse_xlm_file(_file_path):
    try:
        wb_obj = openpyxl.load_workbook(_file_path)
    except FileNotFoundError:
        print(F"{_file_path} not found")
        return {}
    except PermissionError:

        print(F"{_file_path} don't have permission to access this file")
        return {}
    except Exception as error:
        print(error)
        return {}

    try:
        ws = wb_obj.active
    except ValueError:
        return {}
    for i in range(1, ws.max_row + 1):
        data = {
            'no': ws[F'A{i}'].value,
            'employee_id': ws[F'B{i}'].value,
            'age': ws[F'C{i}'].value,
            'salary': ws[F'D{i}'].value
        }
        yield data


def process_xlm_data(_file_path):
    age_wise = {}
    for i in parse_xlm_file(_file_path):
        age = i.get('age')
        salary = i.get('salary')
        employee_id = i.get('employee_id')
        if not isinstance(age, int):
            continue
        if 21 <= age <= 30:
            _age = '21-30'
        elif 31 <= age <= 40:
            _age = '31-40'
        elif 41 <= age <= 51:
            _age = '41-50'
        else:
            _age = '51+'
        if _age not in age_wise:
            age_wise[_age] = dict()

        if salary not in age_wise[_age]:
            age_wise[_age][salary] = list()

        age_wise[_age][salary].append(employee_id)
    return age_wise


def generate_report(data):
    highest_salary = 0
    _message = ''
    for age_group in ['21-30', '31-40', '41-50', '51+']:
        emp_max_salary = max(data[age_group].keys())
        print(F'{age_group}: {emp_max_salary}')
        if highest_salary < emp_max_salary:
            highest_salary = emp_max_salary
            _message = F"{data[age_group][emp_max_salary]} " \
                       F"gets {emp_max_salary} Salary"
    return _message
print('==========================================================')

FILE_NAME = 'test file_python2.xlsx'
FILE_PATH = generate_xls_path(FILE_NAME)
data = process_xlm_data(FILE_PATH)
resp = generate_report(data=data)
print(resp)


# try:
#     print(5/0)
# except ZeroDivisionError:
#     print('Your divider must not be Zero (0)')
# except Exception as error:
#     print('Got an Error', error)
# else:
#     print('survived without an error')